import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.parsers import register
from app.schemas import ParseResult, ParseRow


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _parse_amount(cell) -> float | None:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    s = _cell_str(cell).replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


@register("generic", "xlsx")
def parse_generic_xlsx(buf: BytesIO, bank_code: str) -> ParseResult:
    warnings: list[str] = []
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active
    rows_out: list[ParseRow] = []
    dates: list[str] = []

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return ParseResult(rows=[], warnings=["Empty sheet"], bank_code=bank_code, file_type="xlsx")

    headers = [_cell_str(h).lower() for h in header_row]
    date_idx = next((i for i, h in enumerate(headers) if h in ("date", "fecha", "fecha valor")), None)
    desc_idx = next((i for i, h in enumerate(headers) if h in ("description", "descripcion", "detalle", "concepto")), None)
    amt_idx = next(
        (i for i, h in enumerate(headers) if h in ("amount", "importe", "monto", "debit", "credit")),
        None,
    )

    if date_idx is None or amt_idx is None:
        date_idx, desc_idx, amt_idx = 0, 1, 2
        warnings.append("No header match; assumed columns A=date, B=description, C=amount")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        if date_idx >= len(row) or amt_idx >= len(row):
            continue
        d_raw = row[date_idx]
        amt = _parse_amount(row[amt_idx])
        d = _cell_str(d_raw)
        if len(d) == 10 and d[4] == "-" and d[7] == "-":
            pass
        elif isinstance(d_raw, (datetime, date)):
            d = _cell_str(d_raw)
        else:
            continue
        if amt is None:
            continue
        desc = _cell_str(row[desc_idx]) if desc_idx is not None and desc_idx < len(row) else ""
        rows_out.append(ParseRow(date=d, description=desc, amount=abs(amt), raw={}))
        dates.append(d)

    wb.close()
    period = {"from": min(dates) if dates else None, "to": max(dates) if dates else None}
    return ParseResult(rows=rows_out, warnings=warnings, bank_code=bank_code, file_type="xlsx", period=period)
