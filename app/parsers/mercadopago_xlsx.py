"""Mercado Pago account statement XLSX (same logical layout as semicolon CSV)."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.parsers import register
from app.parsers.mercadopago_common import mp_date_to_iso, parse_argentine_amount
from app.schemas import ParseResult, ParseRow

_TRANSACTION_HEADER = "RELEASE_DATE"


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


@register("mercadopago", "xlsx")
def parse_mercadopago_xlsx(buf: BytesIO, bank_code: str) -> ParseResult:
    warnings: list[str] = []
    rows_out: list[ParseRow] = []
    dates: list[str] = []

    wb = load_workbook(buf, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row_idx: int | None = None
        col_date = col_desc = col_amt = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            cells = [_cell_str(c) for c in row]
            if not any(cells):
                continue
            if cells[0].upper() == _TRANSACTION_HEADER:
                header_row_idx = row_idx
                hdr = [c.upper() for c in cells]
                try:
                    col_date = hdr.index("RELEASE_DATE")
                    col_desc = hdr.index("TRANSACTION_TYPE")
                    col_amt = hdr.index("TRANSACTION_NET_AMOUNT")
                except ValueError:
                    warnings.append("Unexpected Mercado Pago XLSX headers.")
                    return ParseResult(
                        rows=[],
                        warnings=warnings,
                        bank_code=bank_code,
                        file_type="xlsx",
                        currency="ARS",
                    )
                break

        if header_row_idx is None:
            warnings.append("Not a Mercado Pago XLSX: missing RELEASE_DATE header.")
            return ParseResult(
                rows=[],
                warnings=warnings,
                bank_code=bank_code,
                file_type="xlsx",
                currency="ARS",
            )

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row:
                continue
            lst = list(row)
            if col_date >= len(lst) or col_amt >= len(lst):
                continue
            d_raw = lst[col_date]
            if isinstance(d_raw, (datetime, date)):
                d = d_raw.date().isoformat() if isinstance(d_raw, datetime) else d_raw.isoformat()
            else:
                d = mp_date_to_iso(_cell_str(d_raw))
            if d is None:
                continue
            net = parse_argentine_amount(_cell_str(lst[col_amt]) if col_amt < len(lst) else None)
            if net is None:
                continue
            if net >= 0:
                continue
            desc = _cell_str(lst[col_desc]) if col_desc is not None and col_desc < len(lst) else ""
            ref_val = ""
            ref_id_idx = col_desc + 1 if col_desc is not None else None
            if ref_id_idx is not None and ref_id_idx < len(lst):
                ref_val = _cell_str(lst[ref_id_idx])
            rows_out.append(
                ParseRow(
                    date=d,
                    description=desc,
                    amount=abs(net),
                    currency="ARS",
                    raw={"reference_id": ref_val, "source": "mercadopago_xlsx"},
                )
            )
            dates.append(d)

    finally:
        wb.close()

    if not rows_out:
        warnings.append("No expense rows (negative amounts) found in XLSX.")

    period = {"from": min(dates) if dates else None, "to": max(dates) if dates else None}
    return ParseResult(
        rows=rows_out,
        warnings=warnings,
        bank_code=bank_code,
        file_type="xlsx",
        currency="ARS",
        period=period,
    )
